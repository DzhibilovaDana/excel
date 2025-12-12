import os
from tempfile import NamedTemporaryFile

import pandas as pd
from openpyxl import Workbook, load_workbook

from processor import (
    RESULT_COLUMNS,
    apply_artifact1_rows,
    load_dataframe,
    merge_artifact2_sheet,
    select_batch,
    update_dataframe_with_artifact1,
)


def test_select_batch_picks_unprocessed_only():
    df = pd.DataFrame(
        [
            {"L1": "a", RESULT_COLUMNS["trigger"]: "x", RESULT_COLUMNS["executor"]: "y", RESULT_COLUMNS["step_description"]: "z", RESULT_COLUMNS["mirapolis_action"]: "m"},
            {"L1": "b", RESULT_COLUMNS["trigger"]: "", RESULT_COLUMNS["executor"]: "", RESULT_COLUMNS["step_description"]: "", RESULT_COLUMNS["mirapolis_action"]: ""},
            {"L1": "c", RESULT_COLUMNS["trigger"]: "", RESULT_COLUMNS["executor"]: "ok", RESULT_COLUMNS["step_description"]: "", RESULT_COLUMNS["mirapolis_action"]: ""},
        ]
    )
    batch = select_batch(df, batch_size=2)
    assert len(batch) == 2
    assert batch.iloc[0]["L1"] == "b"
    assert batch.iloc[1]["L1"] == "c"


def test_apply_artifact1_rows_writes_into_workbook(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Process"
    ws.append(list(RESULT_COLUMNS.values()))
    for _ in range(3):
        ws.append(["", "", "", ""])
    file_path = tmp_path / "book.xlsx"
    wb.save(file_path)

    artifacts = [
        {
            "row_index": 2,
            "trigger": "start",
            "executor": "owner",
            "step_description": "desc",
            "mirapolis_action": "act",
        }
    ]
    wb_loaded = load_workbook(file_path)
    apply_artifact1_rows(wb_loaded, "Process", artifacts)
    wb_loaded.save(file_path)
    wb2 = load_workbook(file_path)
    ws2 = wb2["Process"]
    assert ws2.cell(row=2, column=1).value == "start"
    assert ws2.cell(row=2, column=2).value == "owner"
    assert ws2.cell(row=2, column=3).value == "desc"
    assert ws2.cell(row=2, column=4).value == "act"


def test_update_dataframe_with_artifact1_marks_done():
    df = pd.DataFrame({RESULT_COLUMNS["trigger"]: ["", ""], RESULT_COLUMNS["executor"]: ["", ""], RESULT_COLUMNS["step_description"]: ["", ""], RESULT_COLUMNS["mirapolis_action"]: ["", ""]})
    update_dataframe_with_artifact1(
        df,
        [
            {
                "row_index": 2,
                "trigger": "t",
                "executor": "e",
                "step_description": "d",
                "mirapolis_action": "a",
            }
        ],
    )
    assert df.loc[0, RESULT_COLUMNS["executor"]] == "e"
    assert df.loc[0, "status"] == "done"


def test_merge_artifact2_sheet_deduplicates(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Process"
    ws.append(["L1"])
    path = tmp_path / "book.xlsx"
    wb.save(path)

    merge_artifact2_sheet(wb, [{"l2_name": "L2", "requirements": {"data_master": ["a", "a"]}}])
    wb.save(path)
    wb2 = load_workbook(path)
    ft = wb2["ФТ Mirapolis"]
    rows = list(ft.iter_rows(min_row=2, values_only=True))
    assert rows == [("L2", "data_master", "a")]

