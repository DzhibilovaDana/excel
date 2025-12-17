# processor.py
import argparse
import json
import logging
import os
import shutil
import time
import base64
from dataclasses import dataclass
from io import BytesIO
from typing import Any, Dict, List, Tuple
import math
import pandas as pd
from google import genai
from google.genai import types
from openpyxl import load_workbook
from openpyxl import Workbook
import base64
from docx import Document
import PyPDF2

# Константы выходных колонок артефакта 1
RESULT_COLUMNS = {
    "trigger": "Инициирующее событие",
    "executor": "Исполнитель",
    "step_description": "Описание шага",
    "mirapolis_action": "Действие в системе",
}

REQUIRED_INPUT_COLS = [
    "L1",
    "L2",
    "L3",
    "L4",
    "L5",
    "Mirapolis L4 (Да/Нет)",
    "Ответ Клиента",
    "Комментарии/уточнения",
    "Комментарии со встреч",
    "Проблема текущего состояния",
    "Предложение по решению /улучшению HCM",
    "Функциональные требования",
]

ALIASES = {
    "Предложение по решению /улучшению HCM": [
        "Предожение по решению /улучшению HCM",  # опечатка
        "Предложение по решению /улучшению HCM \nОфис\nМасс\nПервичные"
    ],
    "Ответ Клиента": [
        "Ответ Агроэко L4 в «Агроэко»\nОфис\nМасс"
    ],
    "Mirapolis L4 (Да/Нет)": ["Mirapolis L4"]
}


ARTIFACT2_SHEET = "ФТ Mirapolis"

from typing import Any, Dict, List, Tuple, Optional

@dataclass
class PipelineConfig:
    input_xlsx: str
    output_xlsx: str
    prompt_file: str
    batch_size: int = 10
    gemini_api_key: str = ""
    gemini_model: str = "gemini-2.0-flash"
    log_dir: str = "logs"
    state_dir: str = "snapshots"
    overwrite_input: bool = False
    extra_files_dir: Optional[str] = None   # <-- новое поле

    @classmethod
    def from_env_and_args(cls) -> "PipelineConfig":
        parser = argparse.ArgumentParser(description="Process Excel through Gemini in batches.")
        parser.add_argument("--input", dest="input_xlsx", required=True, help="Входной XLSX файл")
        parser.add_argument("--output", dest="output_xlsx", required=True, help="Итоговый XLSX файл")
        parser.add_argument("--prompt", dest="prompt_file", required=True, help="Путь к файлу промта")
        parser.add_argument("--batch-size", dest="batch_size", type=int, default=int(os.getenv("BATCH_SIZE", 10)))
        parser.add_argument("--model", dest="gemini_model", default=os.getenv("GEMINI_MODEL", "gemini-2.0-flash"))
        parser.add_argument("--log-dir", dest="log_dir", default=os.getenv("LOG_DIR", "logs"))
        parser.add_argument("--state-dir", dest="state_dir", default=os.getenv("STATE_DIR", "snapshots"))
        parser.add_argument("--api-key", dest="gemini_api_key", default=os.getenv("GEMINI_API_KEY", ""))
        parser.add_argument(
            "--overwrite-input",
            dest="overwrite_input",
            action="store_true",
            default=os.getenv("OVERWRITE_INPUT", "false").lower() in ("1", "true", "yes"),
            help="Если задан, после каждого батча перезаписывать input.xlsx текущим output.xlsx",
        )
        parser.add_argument("--extra-files-dir", dest="extra_files_dir", default=os.getenv("EXTRA_FILES_DIR", ""),
            help="Папка с дополнительными файлами (docx/pdf/txt/иные). Файлы из этой папки будут отправлены модели.")

        args = parser.parse_args()
        api_key = args.gemini_api_key or os.getenv("GEMINI_API_KEY", "")
        if not api_key:
            raise ValueError("GEMINI_API_KEY обязателен (env или --api-key).")

        batch_size = args.batch_size or int(os.getenv("BATCH_SIZE", 10))
        if batch_size <= 0:
            raise ValueError("BATCH_SIZE должен быть > 0")

        return cls(
            input_xlsx=args.input_xlsx,
            output_xlsx=args.output_xlsx,
            prompt_file=args.prompt_file,
            batch_size=batch_size,
            gemini_api_key=api_key,
            gemini_model=args.gemini_model,
            log_dir=args.log_dir,
            state_dir=args.state_dir,
            overwrite_input=args.overwrite_input,
            extra_files_dir=args.extra_files_dir or None,
        )


def ensure_output_workbook(input_path: str, output_path: str) -> None:
    if not os.path.exists(output_path):
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        shutil.copyfile(input_path, output_path)


def get_process_sheet_name(workbook) -> str:
    return "Process" if "Process" in workbook.sheetnames else workbook.sheetnames[0]


def load_dataframe(output_path: str, process_sheet: str) -> pd.DataFrame:
    df = pd.read_excel(output_path, sheet_name=process_sheet)
    df = df.reset_index(drop=True)

    # Убедимся, что все результирующие колонки присутствуют и имеют object dtype,
    # а NaN заменены пустыми строками.
    for col in RESULT_COLUMNS.values():
        if col not in df.columns:
            df[col] = ""
        else:
            # Приводим колонку к object и заполняем NaN пустыми строками
            df[col] = df[col].astype(object).fillna("")

    # status
    df["status"] = df.get("status", "")
    df["status"] = df["status"].astype(object).fillna("")

    return df



def _is_empty_cell(value: Any) -> bool:
    """
    Возвращает True, если значение считается пустым:
    - None
    - numpy.nan / float('nan')
    - пустая строка (после strip)
    """
    if value is None:
        return True
    # pandas / numpy NaN проверяем через math.isnan для чисел, или через pandas.isna
    try:
        # Это сработает для np.nan и для float('nan')
        if isinstance(value, float) and math.isnan(value):
            return True
    except Exception:
        pass
    # pandas uses pd.isna for complex types
    try:
        if pd.isna(value):
            return True
    except Exception:
        pass
    # Finally, empty string after strip
    try:
        if isinstance(value, str) and value.strip() == "":
            return True
    except Exception:
        pass
    return False


def is_row_processed(row: pd.Series) -> bool:
    """
    Строка считается обработанной если:
    - статус == 'done' (регистронезависимо)
    ИЛИ
    - все целевые колонки RESULT_COLUMNS заполнены (т.е. НЕ пусты по _is_empty_cell)
    """
    status = row.get("status", "")
    if isinstance(status, str) and status.strip().lower() == "done":
        return True

    for col in RESULT_COLUMNS.values():
        val = row.get(col, "")
        if _is_empty_cell(val):
            # как только находим пустую целевую колонку — строка НЕ обработана
            return False
    # все целевые колонки непустые — строка обработана
    return True


def select_batch_by_l4(df: pd.DataFrame, batch_size: int = None) -> pd.DataFrame:
    """
    Находит первую непроцессированную строку и выбирает
    все последующие строки, относящиеся к тому же L4 (непрерывный блок).
    """
    # 1. Находим индекс первой необработанной строки
    mask_unprocessed = ~df.apply(is_row_processed, axis=1)
    if not mask_unprocessed.any():
        return df.iloc[0:0]

    first_unprocessed_idx = mask_unprocessed[mask_unprocessed].index[0]
    
    # 2. Получаем значение L4 этой строки
    l4_value = df.at[first_unprocessed_idx, "L4"]
    
    # Если L4 пустой, берем просто batch_size строк (fallback)
    if _is_empty_cell(l4_value):
        return df.loc[first_unprocessed_idx:].head(batch_size or 10)

    # 3. Идем вниз от первой строки и собираем все строки с таким же L4
    # (пока не встретим другой L4 или конец файла)
    indices = []
    l4_str_target = str(l4_value).strip()
    
    # Срез df начиная с найденной строки до конца
    subset = df.loc[first_unprocessed_idx:]
    
    for idx, row in subset.iterrows():
        current_l4 = str(row.get("L4", "")).strip()
        if current_l4 == l4_str_target:
            indices.append(idx)
        else:
            # Как только L4 изменился — останавливаемся. 
            # Это гарантирует, что мы берем только текущий процесс.
            break
            
    return df.loc[indices]



def _normalize(name: str) -> str:
    if not isinstance(name, str):
        return ""
    s = name.lower()
    return "".join(ch for ch in s if ch.isalnum())

def build_header_map(ws):
    """
    Возвращает dict: normalized_name -> (original_name, column_index)
    """
    header_map = {}
    for cell in ws[1]:
        orig = cell.value if cell.value is not None else ""
        norm = _normalize(str(orig))
        header_map[norm] = (orig, cell.column)
    return header_map


def _ensure_columns(batch_df: pd.DataFrame, required_cols: List[str]) -> pd.DataFrame:
    """
    Убедиться, что в batch_df есть все required_cols.
    - Если найдена существующая колонка с похожим именем — скопировать её под каноническим именем.
    - Иначе — создать пустую колонку с этим именем.
    Возвращает изменённый batch_df.
    """
    # Нормализованная карта существующих колонок: norm -> original_name
    norm_map = {}
    for col in batch_df.columns:
        try:
            col_str = str(col)
        except Exception:
            continue
        norm = _normalize_header(col_str)
        if norm:
            norm_map[norm] = col

    for req in required_cols:
        req_norm = _normalize_header(req)
        if req_norm in norm_map:
            orig = norm_map[req_norm]
            if orig != req:
                # копируем колонку под каноническим именем
                batch_df[req] = batch_df[orig]
                logging.info("Сопоставлена колонка '%s' -> '%s'.", orig, req)
            # если orig == req — колонка уже с правильным именем
        else:
            # Либеральный поиск по словам (например: переносы строк, доп слова)
            words = [w for w in "".join(ch if ch.isalnum() else " " for ch in req.lower()).split() if len(w) > 2]
            found = None
            if words:
                for col in batch_df.columns:
                    col_low = str(col).lower()
                    if all(word in col_low for word in words):
                        found = col
                        break
            if found:
                batch_df[req] = batch_df[found]
                logging.info("Либерально сопоставлена колонка '%s' -> '%s'.", found, req)
            else:
                # Создаём пустую колонку, чтобы не падать дальше
                batch_df[req] = ""
                logging.warning("Колонка '%s' не найдена — создана пустая колонка.", req)

    return batch_df


def _normalize_header(name: str) -> str:
    """
    Нормализация имени колонки: приводим к нижнему регистру и оставляем только буквенно-цифровые символы.
    Это позволяет сопоставлять заголовки, игнорируя пробелы, переносы строк, скобки, кавычки и т.п.
    """
    if not isinstance(name, str):
        return ""
    s = name.lower()
    # убираем все не буквенно-цифровые символы
    return "".join(ch for ch in s if ch.isalnum())


def ensure_input_headers_in_workbook(wb, process_sheet: str, required_cols: list, aliases: dict | None = None) -> None:
    """
    Приводит заголовки в workbook к каноническим именам:
    - ищет существующие колонки по нормализованным именам и по алиасам;
    - если находит — переименовывает ячейку заголовка в каноническое имя;
    - если не находит — добавляет новую пустую колонку с каноническим именем в конец.
    Затем workbook нужно сохранить (вызов должен сделать вызывающий код).
    """
    if aliases is None:
        aliases = {}

    ws = wb[process_sheet]

    # Читаем текущие заголовки
    header_cells = list(ws[1])
    norm_map = {}  # norm -> (col_idx, orig_name)
    for cell in header_cells:
        val = cell.value
        if val is None:
            continue
        norm = _normalize_header(str(val))
        if norm:
            norm_map[norm] = (cell.column, val)

    # Функция поиска: по нормализованной форме, затем по алиасам, затем по вхождению ключевых слов
    def find_existing_column(req_name: str):
        req_norm = _normalize_header(req_name)
        # exact norm match
        if req_norm in norm_map:
            return norm_map[req_norm][0], norm_map[req_norm][1]
        # aliases
        if req_name in aliases:
            for alt in aliases[req_name]:
                alt_norm = _normalize_header(alt)
                if alt_norm in norm_map:
                    return norm_map[alt_norm][0], norm_map[alt_norm][1]
        # liberal search: все слова req в кол-во >2 должны присутствовать
        words = [w for w in "".join(ch if ch.isalnum() else " " for ch in req_name.lower()).split() if len(w) > 2]
        if words:
            for norm, (col_idx, orig_name) in norm_map.items():
                low = orig_name.lower()
                if all(word in low for word in words):
                    return col_idx, orig_name
        return None

    # Пройти по требуемым колонкам
    for req in required_cols:
        found = find_existing_column(req)
        if found:
            col_idx, orig_name = found
            # переименовать заголовок, если оно отличается
            if orig_name != req:
                ws.cell(row=1, column=col_idx).value = req
        else:
            # добавить новую колонку в конец
            ws.cell(row=1, column=ws.max_column + 1).value = req


def build_batch_payload_for_l4(batch_df: pd.DataFrame) -> Dict[str, Any]:
    required_cols = [
        "L1", "L2", "L3", "L4", "L5", "Mirapolis L4 (Да/Нет)",
        "Ответ Клиента", "Комментарии/уточнения", "Комментарии со встреч",
        "Проблема текущего состояния", "Предложение по решению /улучшению HCM",
        "Функциональные требования",
    ]
    batch_df = _ensure_columns(batch_df, required_cols)
    if batch_df.empty:
        return {"process_l4_name": "", "first_row_excel_index": None, "rows": []}

    first_df_idx = int(batch_df.index.min())
    first_row_excel_index = first_df_idx + 2  # Excel row index (1-based + header)
    process_l4_name = str(batch_df.at[first_df_idx, "L4"] or "")

    rows = []
    for idx, row in batch_df.iterrows():
        excel_row_index = int(idx) + 2
        rows.append({
            "row_index": excel_row_index,
            "l2_name": row.get("L2", ""),
            "L1": row.get("L1", ""),
            "L2": row.get("L2", ""),
            "L3": row.get("L3", ""),
            "L4": row.get("L4", ""),
            "L5": row.get("L5", ""),
            "Mirapolis_L4": row.get("Mirapolis L4 (Да/Нет)", ""),
            "Ответ_Клиента": row.get("Ответ Клиента", ""),
            "Комментарии_уточнения": row.get("Комментарии/уточнения", ""),
            "Комментарии_со_встреч": row.get("Комментарии со встреч", ""),
            "Проблема_текущего_состояния": row.get("Проблема текущего состояния", ""),
            "Предложение_по_решению": row.get("Предложение по решению /улучшению HCM", ""),
            "Функциональные_требования": row.get("Функциональные требования", ""),
            # мета
            "process_l4_name": process_l4_name,
            "first_row_excel_index": first_row_excel_index,
            "is_first_in_l4": excel_row_index == first_row_excel_index,
        })
    return {
        "process_l4_name": process_l4_name,
        "first_row_excel_index": first_row_excel_index,
        "rows": rows,
    }



class GeminiClient:
    def __init__(self, api_key: str, model_name: str, log_dir: str):
        self.model_name = f"models/{model_name}"
        self.client = genai.Client(api_key=api_key)
        safety_categories = [
            "HARM_CATEGORY_HARASSMENT",
            "HARM_CATEGORY_HATE_SPEECH",
            "HARM_CATEGORY_SEXUALLY_EXPLICIT",
            "HARM_CATEGORY_DANGEROUS_CONTENT",
        ]
        self.config = types.GenerateContentConfig(
            temperature=0.1,
            safety_settings=[types.SafetySetting(category=c, threshold="BLOCK_NONE") for c in safety_categories],
        )
        self.log_dir = log_dir
        os.makedirs(self.log_dir, exist_ok=True)

    def call(self, prompt_text: str, batch_json: List[Dict[str, Any]], batch_idx: int, extra_files: List[Dict[str, Any]] = None) -> Dict[str, Any]:
        system_instruction = (
            "Верни только валидный JSON, без markdown, без пояснений и без дополнительного текста. "
            "Язык значений — русский. Учитывай дополнительные файлы (если они предоставлены) как источники информации. "
            "ВАЖНО: входные данные представляют собой одну группу строк одного процесса L4 и содержат поля "
            "`process_l4_name` (строка), `first_row_excel_index` (целое, индекс первой строки Excel этой группы) и `rows` (массив объектов по строкам). "
            "Требование: верни ровно одну запись в `artifact1_rows` для этой группы. Поле `row_index` в этой записи обязано быть равно `first_row_excel_index`. "
            "Модель должна использовать всю информацию из массива `rows` и все дополнительные файлы при формировании значений полей `trigger`, `executor`, "
            "`step_description`, `mirapolis_action`. НЕ возвращай записи `artifact1_rows` для других row_index в группе — такие записи будут игнорироваться. "
            "Допускается возвращать `artifact2_by_l2` как объект требований по L2 в формате, описанном ниже, если это необходимо."
        )
        contract = {
            "artifact1_rows": [
                {
                    "row_index": 123,
                    "process_l4_name": "...",
                    "trigger": "...",           # Инициирующее событие (строка)
                    "executor": "...",          # Исполнитель (строка)
                    "step_description": "...",  # Описание шага (строка)
                    "mirapolis_action": "..."   # Действие в системе (строка)
                }
            ]
        }

        # Формирование секции дополнительных файлов (оставляем вашу текущую логику)
        files_section = ""
        if extra_files:
            parts = []
            for f in extra_files:
                fname = f.get("filename", "unknown")
                mime = f.get("mime", "unknown")
                if f.get("text"):
                    txt = f["text"]
                    max_len = 75000
                    if len(txt) > max_len:
                        txt = txt[:max_len] + "..."
                    parts.append(f"--- Файл: {fname} (mime: {mime}) ---\n{txt}")
                elif f.get("base64"):
                    parts.append(f"--- Файл (base64): {fname} (mime: {mime}) — base64-данные опущены. ---")
                else:
                    parts.append(f"--- Файл: {fname} (mime: {mime}) — нет текстового содержимого ---")
            files_section = "\n\n".join(parts)

        # Полный промт: системная инструкция, пользовательский промт, описание структуры входных данных, ожидаемый JSON и сами данные
        full_prompt_parts = [
            system_instruction,
            prompt_text,
            "Структура входных данных (объект JSON):",
            # тут показываем точную структуру (для модели), но это служебная информация — можно оставить шаблон
            json.dumps({"process_l4_name": "STRING", "first_row_excel_index": 123, "rows": [{"row_index": 123, "L1": "...", "L2": "...", "...": "..."}]}, ensure_ascii=False, indent=2),
            "Ожидаемый JSON (строгая схема):",
            json.dumps(contract, ensure_ascii=False, indent=2),
            "Вот данные (первичный источник, объект):",
            json.dumps(batch_json, ensure_ascii=False, indent=2),
        ]
        full_prompt = "\n\n".join(full_prompt_parts)

        if files_section:
            full_prompt += "\n\n--- Дополнительные файлы для изучения: ---\n" + files_section

        # Логи
        input_log = os.path.join(self.log_dir, f"batch_{batch_idx}_input.json")
        raw_log = os.path.join(self.log_dir, f"batch_{batch_idx}_raw.txt")
        parsed_log = os.path.join(self.log_dir, f"batch_{batch_idx}_parsed.json")

        with open(input_log, "w", encoding="utf-8") as f:
            json.dump(batch_json, f, ensure_ascii=False, indent=2)

        for attempt in range(3):
            try:
                response = self.client.models.generate_content(model=self.model_name, contents=[full_prompt], config=self.config)

                # --- robust extraction (ваша существующая логика) ---
                response_text = None

                if hasattr(response, "text"):
                    r = getattr(response, "text")
                    if isinstance(r, str):
                        response_text = r.strip()
                    else:
                        try:
                            response_text = json.dumps(r, ensure_ascii=False)
                        except Exception:
                            response_text = str(r)

                if response_text is None:
                    cand = getattr(response, "candidates", None)
                    if cand and isinstance(cand, (list, tuple)) and len(cand) > 0:
                        first = cand[0]
                        if isinstance(first, dict):
                            if "content" in first and isinstance(first["content"], (list, str)):
                                response_text = first["content"][0] if isinstance(first["content"], list) else first["content"]
                            elif "text" in first:
                                response_text = first["text"]
                            else:
                                response_text = json.dumps(first, ensure_ascii=False)
                        else:
                            cont = getattr(first, "content", None)
                            if cont and isinstance(cont, (list, tuple)) and len(cont) > 0:
                                elem = cont[0]
                                if isinstance(elem, dict):
                                    response_text = elem.get("text") or json.dumps(elem, ensure_ascii=False)
                                else:
                                    response_text = getattr(elem, "text", str(elem))
                            else:
                                response_text = str(first)

                if response_text is None:
                    try:
                        response_text = str(response)
                    except Exception:
                        response_text = ""

                response_text = (response_text or "").strip()
                with open(raw_log, "w", encoding="utf-8") as f:
                    f.write(response_text)

                # Очищаем backticks и префиксы
                cleaned = response_text
                if cleaned.startswith("```"):
                    cleaned = cleaned.strip("`").strip()
                    if cleaned.lower().startswith("json"):
                        cleaned = cleaned[4:].strip()

                parsed = None
                try:
                    parsed = json.loads(cleaned)
                except Exception:
                    parsed = None

                # Если parsed - список — привести к ожидаемой структуре
                if isinstance(parsed, list):
                    logging.warning("Gemini вернул JSON-массив (list). Попытка привести к ожидаемой структуре.")
                    if all(isinstance(el, dict) and "row_index" in el for el in parsed):
                        parsed = {"artifact1_rows": parsed}
                    elif all(isinstance(el, dict) and "l2_name" in el for el in parsed):
                        parsed = {"artifact2_by_l2": parsed}
                    else:
                        parsed = {"raw_list": parsed}

                # Если parsed — словарь и содержит base64 xlsx
                if isinstance(parsed, dict) and "filled_xlsx_base64" in parsed:
                    try:
                        b64 = parsed["filled_xlsx_base64"]
                        xbytes = base64.b64decode(b64)
                        filled_path = os.path.join(self.log_dir, f"batch_{batch_idx}_filled_from_gemini.xlsx")
                        wb = load_workbook(filename=BytesIO(xbytes))
                        wb.save(filled_path)
                        parsed_no_b64 = dict(parsed)
                        parsed_no_b64.pop("filled_xlsx_base64", None)
                        parsed_no_b64["_filled_xlsx_path"] = filled_path
                        with open(parsed_log, "w", encoding="utf-8") as f:
                            json.dump(parsed_no_b64, f, ensure_ascii=False, indent=2)
                        return {"filled_xlsx_path": filled_path, **parsed_no_b64}
                    except Exception as exc:
                        logging.warning("Ошибка декодирования base64 xlsx: %s", exc)
                        raise

                if isinstance(parsed, dict):
                    logging.debug("Parsed gemini json keys: %s", list(parsed.keys()))
                    with open(parsed_log, "w", encoding="utf-8") as f:
                        json.dump(parsed, f, ensure_ascii=False, indent=2)
                    return parsed

                # Попытка распознать base64 xlsx в raw тексте
                text = cleaned.strip()
                if len(text) > 300 and all(c in "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789+/=\r\n" for c in text[:200]):
                    try:
                        xbytes = base64.b64decode(text)
                        filled_path = os.path.join(self.log_dir, f"batch_{batch_idx}_filled_from_gemini.xlsx")
                        wb = load_workbook(filename=BytesIO(xbytes))
                        wb.save(filled_path)
                        with open(parsed_log, "w", encoding="utf-8") as f:
                            json.dump({"filled_xlsx_path": filled_path}, f, ensure_ascii=False, indent=2)
                        return {"filled_xlsx_path": filled_path}
                    except Exception:
                        pass

                with open(parsed_log, "w", encoding="utf-8") as f:
                    json.dump({"raw_text": response_text}, f, ensure_ascii=False, indent=2)
                raise RuntimeError("Не удалось получить валидный JSON/файл от Gemini после разбора.")

            except Exception as exc:
                logging.warning("Gemini ошибка, попытка %s: %s", attempt + 1, exc)
                time.sleep(15)

        raise RuntimeError("Не удалось получить валидный JSON/файл от Gemini после 3 попыток.")




def ensure_result_columns(wb, process_sheet: str) -> None:
    ws = wb[process_sheet]
    # Убедимся, что заголовки существуют в нужном порядке (но не портим существующие):
    headers = [cell.value for cell in ws[1]]
    # Используем нормализованные имена для поиска
    header_map = { _normalize(h or ""): (h, idx+1) for idx, h in enumerate(headers) }
    for col_name in RESULT_COLUMNS.values():
        norm = _normalize(col_name)
        if norm not in header_map:
            # добавляем в конец
            ws.cell(row=1, column=ws.max_column + 1).value = col_name
            logging.info("Добавлена результирующая колонка '%s' в sheet %s", col_name, process_sheet)


def apply_artifact1_rows(wb, process_sheet: str, artifact_rows: List[Dict[str, Any]], allowed_first_row_index: int = None) -> None:
    ws = wb[process_sheet]
    header_map = build_header_map(ws)
    # Для быстрых обращений: норм -> column index (если нет — создадим)
    for canonical in RESULT_COLUMNS.values():
        norm = _normalize(canonical)
        if norm not in header_map:
            ws.cell(row=1, column=ws.max_column + 1).value = canonical
            header_map[norm] = (canonical, ws.max_column)
            logging.info("Создана колонка результатов '%s' в листе '%s'", canonical, process_sheet)

    # применяем
    for item in artifact_rows:
        try:
            row_idx = int(item["row_index"])
        except Exception:
            logging.warning("Некорректный row_index в артефакте: %s", item.get("row_index"))
            continue

        # Если указан allowed_first_row_index — пропускаем записи для других индексов
        if allowed_first_row_index is not None and row_idx != int(allowed_first_row_index):
            logging.info("Пропускаю запись artifact1 для row %s — не первая строка L4 (allowed=%s).", row_idx, allowed_first_row_index)
            continue

        def write_col(key_name, value):
            target_col_name = RESULT_COLUMNS[key_name]
            norm = _normalize(target_col_name)
            col_idx = header_map[norm][1]
            ws.cell(row=row_idx, column=col_idx).value = value or ""

        write_col("trigger", item.get("trigger", ""))
        write_col("executor", item.get("executor", ""))
        write_col("step_description", item.get("step_description", ""))
        write_col("mirapolis_action", item.get("mirapolis_action", ""))


def merge_artifact2_sheet(wb, artifact2: List[Dict[str, Any]]) -> None:
    if ARTIFACT2_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(ARTIFACT2_SHEET)
        ws.append(["L2", "Категория", "Требование"])
    ws = wb[ARTIFACT2_SHEET]
    existing = set()
    for l2, category, req in ws.iter_rows(min_row=2, values_only=True):
        existing.add((l2, category, req))

    new_rows: List[Tuple[str, str, str]] = []
    for item in artifact2:
        l2 = item.get("l2_name", "")
        reqs = item.get("requirements", {})
        for category, items in reqs.items():
            for text in items or []:
                candidate = (l2, category, text)
                if candidate not in existing:
                    existing.add(candidate)
                    new_rows.append(candidate)
    for row in new_rows:
        ws.append(list(row))


def merge_workbooks_preserve(old_wb: Workbook, new_wb: Workbook, process_sheet_name: str):
    """
    Переносит непустые значения из new_wb в old_wb для sheet process_sheet_name.
    Добавляет колонки из new_wb, если их нет в old_wb.
    Объединяет ARTIFACT2_SHEET с дедупом.
    """
    if process_sheet_name not in new_wb.sheetnames:
        return

    old_ws = old_wb[process_sheet_name]
    new_ws = new_wb[process_sheet_name]

    header_to_col_old = {cell.value: cell.column for cell in old_ws[1]}
    header_to_col_new = {cell.value: cell.column for cell in new_ws[1]}

    # Добавляем отсутствующие колонки
    for col_name in header_to_col_new.keys():
        if col_name not in header_to_col_old:
            old_ws.cell(row=1, column=old_ws.max_column + 1).value = col_name
            header_to_col_old[col_name] = old_ws.max_column

    # Перенос непустых значений
    for r in range(2, new_ws.max_row + 1):
        for col_name, new_col_idx in header_to_col_new.items():
            new_val = new_ws.cell(row=r, column=new_col_idx).value
            if new_val is not None and str(new_val).strip() != "":
                old_col_idx = header_to_col_old[col_name]
                old_ws.cell(row=r, column=old_col_idx).value = new_val

    # Объединение ФТ Mirapolis
    if ARTIFACT2_SHEET in new_wb.sheetnames:
        if ARTIFACT2_SHEET not in old_wb.sheetnames:
            old_ws2 = old_wb.create_sheet(ARTIFACT2_SHEET)
            old_ws2.append(["L2", "Категория", "Требование"])
        else:
            old_ws2 = old_wb[ARTIFACT2_SHEET]
        new_ws2 = new_wb[ARTIFACT2_SHEET]

        existing = set()
        for l2, cat, req in old_ws2.iter_rows(min_row=2, values_only=True):
            existing.add(((l2 or ""), (cat or ""), (req or "")))

        for r in range(2, new_ws2.max_row + 1):
            l2 = new_ws2.cell(row=r, column=1).value or ""
            cat = new_ws2.cell(row=r, column=2).value or ""
            req = new_ws2.cell(row=r, column=3).value or ""
            candidate = (l2, cat, req)
            if candidate not in existing:
                old_ws2.append(list(candidate))
                existing.add(candidate)


def update_dataframe_with_artifact1(df: pd.DataFrame, artifact_rows: List[Dict[str, Any]]) -> None:
    for item in artifact_rows:
        try:
            excel_row = int(item["row_index"])
        except Exception:
            logging.warning("Некорректный row_index в artifact_rows: %s", item.get("row_index"))
            continue
        # ожидаем, что строка в df соответствует excel_row - 2 (первый data row = excel row 2)
        df_idx = excel_row - 2
        if df_idx < 0 or df_idx >= len(df):
            logging.warning("row_index %s вне диапазона data frame (df length %s). Пропускаю запись в df.", excel_row, len(df))
            continue
        # если колонок нет в df — добавляем
        for key in RESULT_COLUMNS.values():
            if key not in df.columns:
                df[key] = ""
        df.at[df_idx, RESULT_COLUMNS["trigger"]] = item.get("trigger", "")
        df.at[df_idx, RESULT_COLUMNS["executor"]] = item.get("executor", "")
        df.at[df_idx, RESULT_COLUMNS["step_description"]] = item.get("step_description", "")
        df.at[df_idx, RESULT_COLUMNS["mirapolis_action"]] = item.get("mirapolis_action", "")
        df.at[df_idx, "status"] = "done"



def process_excel(cfg: PipelineConfig) -> None:
    logging.info("Запуск обработки: %s", cfg)
    os.makedirs(cfg.log_dir, exist_ok=True)
    os.makedirs(cfg.state_dir, exist_ok=True)
    ensure_output_workbook(cfg.input_xlsx, cfg.output_xlsx)

    wb = load_workbook(cfg.output_xlsx)
    process_sheet = get_process_sheet_name(wb)
    ensure_result_columns(wb, process_sheet)
    ensure_input_headers_in_workbook(wb, process_sheet, REQUIRED_INPUT_COLS, aliases=ALIASES)
    wb.save(cfg.output_xlsx)
    df = load_dataframe(cfg.output_xlsx, process_sheet)

    batch_idx = 1
    client = GeminiClient(cfg.gemini_api_key, cfg.gemini_model, cfg.log_dir)

    # ----------------- Подготовка дополнительных файлов -----------------
    extra_files_payload = []
    if cfg.extra_files_dir:
        extra_dir = cfg.extra_files_dir
        if os.path.isdir(extra_dir):
            logging.info("Загружаю дополнительные файлы из: %s", extra_dir)
            files = sorted(os.listdir(extra_dir))
            for fname in files:
                fpath = os.path.join(extra_dir, fname)
                if not os.path.isfile(fpath):
                    continue
                ext = os.path.splitext(fname)[1].lower()
                try:
                    # DOCX
                    if ext == ".docx":
                        try:
                            from docx import Document
                            doc = Document(fpath)
                            text = "\n".join([p.text for p in doc.paragraphs if p.text])
                            extra_files_payload.append({"filename": fname, "mime": "application/vnd.openxmlformats-officedocument.wordprocessingml.document", "text": text})
                        except Exception as e:
                            logging.warning("Не удалось прочитать docx %s: %s. Отправлю как base64.", fpath, e)
                            with open(fpath, "rb") as ff:
                                b64 = base64.b64encode(ff.read()).decode("ascii")
                            extra_files_payload.append({"filename": fname, "mime": "application/octet-stream", "base64": b64})
                    # PDF
                    elif ext == ".pdf":
                        try:
                            import PyPDF2
                            text = ""
                            with open(fpath, "rb") as ff:
                                reader = PyPDF2.PdfReader(ff)
                                for page in reader.pages:
                                    text += "\n" + (page.extract_text() or "")
                            extra_files_payload.append({"filename": fname, "mime": "application/pdf", "text": text})
                        except Exception as e:
                            logging.warning("Не удалось прочитать pdf %s: %s. Отправлю как base64.", fpath, e)
                            with open(fpath, "rb") as ff:
                                b64 = base64.b64encode(ff.read()).decode("ascii")
                            extra_files_payload.append({"filename": fname, "mime": "application/octet-stream", "base64": b64})
                    # TXT / MD
                    elif ext in (".txt", ".md"):
                        try:
                            with open(fpath, "r", encoding="utf-8") as ff:
                                text = ff.read()
                            extra_files_payload.append({"filename": fname, "mime": "text/plain", "text": text})
                        except Exception as e:
                            logging.warning("Ошибка чтения текстового файла %s: %s. Отправлю как base64.", fpath, e)
                            with open(fpath, "rb") as ff:
                                b64 = base64.b64encode(ff.read()).decode("ascii")
                            extra_files_payload.append({"filename": fname, "mime": "application/octet-stream", "base64": b64})
                    # XLSX/XLS и прочие бинарные — отправляем base64
                    else:
                        with open(fpath, "rb") as ff:
                            b64 = base64.b64encode(ff.read()).decode("ascii")
                        extra_files_payload.append({"filename": fname, "mime": "application/octet-stream", "base64": b64})
                    logging.info("Подготовлен файл %s (ext=%s).", fname, ext)
                except Exception as exc:
                    logging.warning("Ошибка при обработке доп. файла %s: %s", fpath, exc)
        else:
            logging.warning("Папка с дополнительными файлами не существует: %s", extra_dir)
    else:
        logging.info("Дополнительная папка файлов не указана.")
    # -------------------------------------------------------------------

    while True:
        # 1. Выбираем батч (один процесс L4 целиком)
        batch_df = select_batch_by_l4(df, cfg.batch_size)
        if batch_df.empty:
            logging.info("Необработанных строк нет.")
            break

        logging.info("Обработка батча %d. L4='%s'. Строк: %d (Excel row %d)", 
                     batch_idx, 
                     batch_df.iloc[0].get("L4"), 
                     len(batch_df),
                     int(batch_df.index[0]) + 2)

        # собираем payload
        batch_payload = build_batch_payload_for_l4(batch_df)

        with open(cfg.prompt_file, "r", encoding="utf-8") as f:
            prompt_text = f.read()

        # CSV лог
        pre_batch_csv = os.path.join(cfg.log_dir, f"batch_{batch_idx}_pre.csv")
        batch_df.to_csv(pre_batch_csv, index=False, encoding="utf-8-sig")

        # Вызов API
        try:
            response = client.call(prompt_text, batch_payload, batch_idx, extra_files=extra_files_payload)
        except Exception as e:
            logging.error("Критическая ошибка при вызове Gemini для батча %s: %s", batch_idx, e)
            # Можно прервать или пропустить, но лучше break, чтобы не крутить бесконечно
            break

        first_row_idx = batch_payload.get("first_row_excel_index")

        # --- ОБРАБОТКА ОТВЕТА ---
        if isinstance(response, dict) and response.get("filled_xlsx_path"):
            # Если вернулся файл
            filled_path = response["filled_xlsx_path"]
            logging.info("Gemini вернул готовый XLSX. Merge...")
            new_wb = load_workbook(filled_path)
            merge_workbooks_preserve(wb, new_wb, process_sheet)
            ensure_input_headers_in_workbook(wb, process_sheet, REQUIRED_INPUT_COLS, aliases=ALIASES)
            
            # Обновляем df и сохраняем
            wb.save(cfg.output_xlsx)
            df = load_dataframe(cfg.output_xlsx, process_sheet) # Перечитываем статусы
            
        else:
            # Если вернулся JSON
            artifact1_rows = response.get("artifact1_rows", [])
            artifact2_by_l2 = response.get("artifact2_by_l2", [])
            
            # Фильтруем (на всякий случай, хотя промт запрещает лишние)
            if first_row_idx is not None:
                artifact1_rows = [r for r in artifact1_rows if int(r.get("row_index", -1)) == int(first_row_idx)]
            
            if artifact1_rows:
                apply_artifact1_rows(wb, process_sheet, artifact1_rows, allowed_first_row_index=first_row_idx)
                update_dataframe_with_artifact1(df, artifact1_rows)
            
            if artifact2_by_l2:
                merge_artifact2_sheet(wb, artifact2_by_l2)

            # --- ВАЖНОЕ ИСПРАВЛЕНИЕ: ПОМЕЧАЕМ ВЕСЬ БАТЧ КАК "DONE" ---
            # Даже если мы записали ответ только в 1-ю строку, 
            # мы считаем, что ВЕСЬ L4 процесс обработан.
            # Иначе на следующем шаге select_batch_by_l4 возьмет 2-ю строку этого же L4.
            for idx in batch_df.index:
                df.at[idx, "status"] = "done"

            wb.save(cfg.output_xlsx)
            snapshot_path = os.path.join(cfg.state_dir, f"output_batch_{batch_idx}.xlsx")
            wb.save(snapshot_path)
            logging.info("Батч %s обработан. Снапшот: %s", batch_idx, snapshot_path)

        # Сохранение и перезапись input (как было у тебя)
        post_batch_csv = os.path.join(cfg.log_dir, f"batch_{batch_idx}_post.csv")
        df.to_csv(post_batch_csv, index=False, encoding="utf-8-sig")

        if cfg.overwrite_input:
            try:
                if os.path.abspath(cfg.input_xlsx) != os.path.abspath(cfg.output_xlsx):
                    shutil.copyfile(cfg.output_xlsx, cfg.input_xlsx)
            except Exception: pass

        batch_idx += 1


def main():
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
    cfg = PipelineConfig.from_env_and_args()
    process_excel(cfg)


if __name__ == "__main__":
    main()
